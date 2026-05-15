import frappe
import io
from datetime import date, timedelta


# ── Defaults ──────────────────────────────────────────────────────────────────

DEFAULT_CATEGORY_MAP = {
    'Watches':     'Watches',
    'Watch Service': 'Services',
    'Jewelry':     'Jewellery',
    'Jewellery':   'Jewellery',
    'Accessories': 'Accessories',
    'Strap':       'Accessories',
}


# ── Config loader ─────────────────────────────────────────────────────────────

def get_config():
    try:
        doc     = frappe.get_single('AOT Dashboard Config')
        cat_map = {r.item_group: r.category_label for r in (doc.category_mapping or [])}
        hv_ids  = {r.customer for r in (doc.high_end_customers or []) if r.customer}
        ex_ids  = {r.customer for r in (doc.removed_customers  or []) if r.customer}
        return {
            'category_map':      cat_map or DEFAULT_CATEGORY_MAP,
            'hv_customers':      hv_ids,
            'exclude_customers': ex_ids,
        }
    except Exception:
        return {
            'category_map':      DEFAULT_CATEGORY_MAP,
            'hv_customers':      set(),
            'exclude_customers': set(),
        }


# ── Period date ranges ────────────────────────────────────────────────────────

def _safe_ly(d):
    try:
        return d.replace(year=d.year - 1)
    except ValueError:
        return d.replace(year=d.year - 1, day=28)


def get_periods():
    today = date.today()

    # Weekly: last complete week (Mon–Sun) vs the week before it
    last_sun      = today - timedelta(days=today.weekday() + 1)
    last_mon      = last_sun  - timedelta(days=6)
    prev_week_sun = last_mon  - timedelta(days=1)
    prev_week_mon = prev_week_sun - timedelta(days=6)

    # MTD
    mtd_start = today.replace(day=1)

    # Indian FY
    fy_year = today.year if today.month >= 4 else today.year - 1

    # QTD
    m = today.month
    if m in (4, 5, 6):   q_start = date(fy_year, 4, 1)
    elif m in (7, 8, 9): q_start = date(fy_year, 7, 1)
    elif m in (10,11,12):q_start = date(fy_year, 10, 1)
    else:                 q_start = date(today.year, 1, 1)

    ytd_start = date(fy_year, 4, 1)

    def fmt(d):   return d.strftime('%-d %b %Y')
    def short(d): return d.strftime('%-d %b\'%y')

    cy_yr = today.year
    ly_yr = today.year - 1

    q_labels = {4: 'Q1', 7: 'Q2', 10: 'Q3', 1: 'Q4'}
    q_label  = q_labels.get(q_start.month, 'Q')

    def wk_col(s, e):
        return short(e) if s == e else f"{s.strftime('%-d')}–{short(e)}"

    return {
        'daily': {
            'label':    'Daily',
            'cy_start': today,           'cy_end': today,
            'ly_start': _safe_ly(today), 'ly_end': _safe_ly(today),
            'cy_label': fmt(today),
            'ly_label': fmt(_safe_ly(today)),
            'cy_col':   short(today),
            'ly_col':   short(_safe_ly(today)),
        },
        'weekly': {
            'label':    'Weekly',
            'cy_start': last_mon,      'cy_end': last_sun,
            'ly_start': prev_week_mon, 'ly_end': prev_week_sun,
            'cy_label': f'{fmt(last_mon)} – {fmt(last_sun)}',
            'ly_label': f'{fmt(prev_week_mon)} – {fmt(prev_week_sun)}',
            'cy_col':   wk_col(last_mon, last_sun),
            'ly_col':   wk_col(prev_week_mon, prev_week_sun),
        },
        'mtd': {
            'label':    'MTD',
            'cy_start': mtd_start,           'cy_end': today,
            'ly_start': _safe_ly(mtd_start), 'ly_end': _safe_ly(today),
            'cy_label': f'{fmt(mtd_start)} – {fmt(today)}',
            'ly_label': f'{fmt(_safe_ly(mtd_start))} – {fmt(_safe_ly(today))}',
            'cy_col':   today.strftime('%b\'%y'),
            'ly_col':   _safe_ly(today).strftime('%b\'%y'),
        },
        'qtd': {
            'label':    'QTD',
            'cy_start': q_start,           'cy_end': today,
            'ly_start': _safe_ly(q_start), 'ly_end': _safe_ly(today),
            'cy_label': f'{fmt(q_start)} – {fmt(today)}',
            'ly_label': f'{fmt(_safe_ly(q_start))} – {fmt(_safe_ly(today))}',
            'cy_col':   f'{q_label} FY{str(cy_yr)[2:]}',
            'ly_col':   f'{q_label} FY{str(ly_yr)[2:]}',
        },
        'ytd': {
            'label':    'YTD',
            'cy_start': ytd_start,           'cy_end': today,
            'ly_start': _safe_ly(ytd_start), 'ly_end': _safe_ly(today),
            'cy_label': f'{fmt(ytd_start)} – {fmt(today)}',
            'ly_label': f'{fmt(_safe_ly(ytd_start))} – {fmt(_safe_ly(today))}',
            'cy_col':   f'FY {cy_yr}-{str(cy_yr + 1)[2:]}',
            'ly_col':   f'FY {ly_yr}-{str(ly_yr + 1)[2:]}',
        },
    }


# ── SQL fetch ─────────────────────────────────────────────────────────────────

def fetch_rows(start_date, end_date, config):
    raw = frappe.db.sql("""
        SELECT
            si.customer,
            si.customer_name,
            sii.item_group,
            COALESCE(sii.brand, '')       AS brand,
            COALESCE(sii.cost_center, '') AS cost_center,
            SUM(sii.base_net_amount) AS amount,
            SUM(sii.qty)             AS qty
        FROM `tabSales Invoice` si
        JOIN `tabSales Invoice Item` sii ON sii.parent = si.name
        WHERE si.docstatus = 1
          AND si.is_return  = 0
          AND si.posting_date BETWEEN %(start_date)s AND %(end_date)s
        GROUP BY si.customer, sii.item_group, sii.brand, sii.cost_center
        ORDER BY si.customer
    """, {'start_date': start_date, 'end_date': end_date}, as_dict=True)

    return _tag(raw, config)


def _tag(rows, config):
    tagged  = []
    hv_ids  = config['hv_customers']
    ex_ids  = config['exclude_customers']
    cat_map = config['category_map']

    for row in rows:
        customer_id = row.get('customer') or ''
        if customer_id in ex_ids:
            continue
        category = cat_map.get(row.get('item_group') or '')
        if not category:
            continue
        ctype = 'High Value' if customer_id in hv_ids else 'B2C'
        tagged.append({
            'customer':      customer_id,
            'customer_name': row.get('customer_name') or customer_id,
            'item_group':    row['item_group'],
            'category':      category,
            'customer_type': ctype,
            'brand':         row.get('brand') or '',
            'cost_center':   row.get('cost_center') or '',
            'amount':        float(row.get('amount') or 0),
            'qty':           float(row.get('qty') or 0),
        })
    return tagged


# ── Aggregation helpers ───────────────────────────────────────────────────────

def _sum(rows, ctype=None, category=None):
    amt = qty = 0.0
    for r in rows:
        if ctype    and r['customer_type'] != ctype:    continue
        if category and r['category']      != category: continue
        amt += r['amount']
        qty += r['qty']
    return amt, qty


def _growth(cy, ly):
    if ly and ly != 0:
        return round((cy - ly) / abs(ly), 4)
    return None


def _row(key, label, indent, cy_amt, cy_qty, ly_amt, ly_qty):
    return {
        'key':        key,
        'label':      label,
        'indent':     indent,
        'cy_rev':     round(cy_amt, 2),
        'ly_rev':     round(ly_amt, 2),
        'growth_rev': _growth(cy_amt, ly_amt),
        'cy_qty':     int(round(cy_qty)),
        'ly_qty':     int(round(ly_qty)),
        'growth_qty': _growth(cy_qty, ly_qty),
    }


def aggregate_snapshot(cy, ly):
    cy_net,   cy_net_q   = _sum(cy)
    ly_net,   ly_net_q   = _sum(ly)
    cy_b2c,   cy_b2c_q   = _sum(cy, 'B2C')
    ly_b2c,   ly_b2c_q   = _sum(ly, 'B2C')
    cy_b2c_w, cy_b2c_wq  = _sum(cy, 'B2C', 'Watches')
    ly_b2c_w, ly_b2c_wq  = _sum(ly, 'B2C', 'Watches')
    cy_b2c_a, cy_b2c_aq  = _sum(cy, 'B2C', 'Accessories')
    ly_b2c_a, ly_b2c_aq  = _sum(ly, 'B2C', 'Accessories')
    cy_b2c_j, cy_b2c_jq  = _sum(cy, 'B2C', 'Jewellery')
    ly_b2c_j, ly_b2c_jq  = _sum(ly, 'B2C', 'Jewellery')
    cy_b2c_s, cy_b2c_sq  = _sum(cy, 'B2C', 'Services')
    ly_b2c_s, ly_b2c_sq  = _sum(ly, 'B2C', 'Services')
    cy_hv,    cy_hv_q    = _sum(cy, 'High Value')
    ly_hv,    ly_hv_q    = _sum(ly, 'High Value')
    cy_hv_w,  cy_hv_wq   = _sum(cy, 'High Value', 'Watches')
    ly_hv_w,  ly_hv_wq   = _sum(ly, 'High Value', 'Watches')
    cy_hv_a,  cy_hv_aq   = _sum(cy, 'High Value', 'Accessories')
    ly_hv_a,  ly_hv_aq   = _sum(ly, 'High Value', 'Accessories')
    cy_hv_j,  cy_hv_jq   = _sum(cy, 'High Value', 'Jewellery')
    ly_hv_j,  ly_hv_jq   = _sum(ly, 'High Value', 'Jewellery')
    cy_hv_s,  cy_hv_sq   = _sum(cy, 'High Value', 'Services')
    ly_hv_s,  ly_hv_sq   = _sum(ly, 'High Value', 'Services')

    return [
        _row('net_revenue',    'Net Revenue', 0, cy_net,   cy_net_q,   ly_net,   ly_net_q),
        _row('b2c',            'B2C',         0, cy_b2c,   cy_b2c_q,   ly_b2c,   ly_b2c_q),
        _row('b2c_watches',    'Watches',     1, cy_b2c_w, cy_b2c_wq,  ly_b2c_w, ly_b2c_wq),
        _row('b2c_accessories','Accessories', 1, cy_b2c_a, cy_b2c_aq,  ly_b2c_a, ly_b2c_aq),
        _row('b2c_jewellery',  'Jewellery',   1, cy_b2c_j, cy_b2c_jq,  ly_b2c_j, ly_b2c_jq),
        _row('b2c_services',   'Services',    1, cy_b2c_s, cy_b2c_sq,  ly_b2c_s, ly_b2c_sq),
        _row('high_value',     'High Value',  0, cy_hv,    cy_hv_q,    ly_hv,    ly_hv_q),
        _row('hv_watches',     'Watches',     1, cy_hv_w,  cy_hv_wq,   ly_hv_w,  ly_hv_wq),
        _row('hv_accessories', 'Accessories', 1, cy_hv_a,  cy_hv_aq,   ly_hv_a,  ly_hv_aq),
        _row('hv_jewellery',   'Jewellery',   1, cy_hv_j,  cy_hv_jq,   ly_hv_j,  ly_hv_jq),
        _row('hv_services',    'Services',    1, cy_hv_s,  cy_hv_sq,   ly_hv_s,  ly_hv_sq),
    ]


def _aggregate_perf(cy, ly, key_field, b2c_only=True):
    """Aggregate by brand or cost_center. b2c_only=True excludes High Value rows."""
    cy_b2c = [r for r in cy if (not b2c_only or r['customer_type'] == 'B2C') and r.get(key_field)]
    ly_b2c = [r for r in ly if (not b2c_only or r['customer_type'] == 'B2C') and r.get(key_field)]

    all_keys = sorted(
        {r[key_field] for r in cy_b2c} | {r[key_field] for r in ly_b2c}
    )
    result = []
    for k in all_keys:
        cy_r = [r for r in cy_b2c if r[key_field] == k]
        ly_r = [r for r in ly_b2c if r[key_field] == k]

        cy_amt = sum(r['amount'] for r in cy_r)
        cy_qty = sum(r['qty']    for r in cy_r)
        ly_amt = sum(r['amount'] for r in ly_r)
        ly_qty = sum(r['qty']    for r in ly_r)
        cy_asp = cy_amt / cy_qty if cy_qty else 0.0
        ly_asp = ly_amt / ly_qty if ly_qty else 0.0

        result.append({
            'name':         k,
            'cy_sales':     round(cy_amt, 2),
            'ly_sales':     round(ly_amt, 2),
            'growth_sales': _growth(cy_amt, ly_amt),
            'cy_units':     int(round(cy_qty)),
            'ly_units':     int(round(ly_qty)),
            'growth_units': _growth(cy_qty, ly_qty),
            'cy_asp':       round(cy_asp, 2),
            'ly_asp':       round(ly_asp, 2),
            'growth_asp':   _growth(cy_asp, ly_asp),
        })

    result.sort(key=lambda x: x['cy_sales'], reverse=True)
    return result


# ── Debug helper ─────────────────────────────────────────────────────────────

@frappe.whitelist()
def debug_category_breakdown(start_date=None, end_date=None):
    """
    Returns raw per-item_group totals for a date range so you can
    cross-check against the item-wise sales register.
    Call from browser console:
      frappe.call({ method: '...aot_data.debug_category_breakdown',
                    args: { start_date: '2026-05-15', end_date: '2026-05-15' },
                    callback: r => console.table(r.message) })
    """
    if not start_date:
        start_date = end_date = str(date.today())
    if not end_date:
        end_date = start_date

    rows = frappe.db.sql("""
        SELECT
            sii.item_group,
            SUM(sii.base_net_amount) AS base_net_amount,
            SUM(sii.net_amount)      AS net_amount,
            SUM(sii.amount)          AS amount,
            SUM(sii.qty)             AS qty,
            COUNT(DISTINCT si.name)  AS invoices
        FROM `tabSales Invoice` si
        JOIN `tabSales Invoice Item` sii ON sii.parent = si.name
        WHERE si.docstatus = 1
          AND si.is_return  = 0
          AND si.posting_date BETWEEN %(start_date)s AND %(end_date)s
        GROUP BY sii.item_group
        ORDER BY base_net_amount DESC
    """, {'start_date': start_date, 'end_date': end_date}, as_dict=True)

    config  = get_config()
    cat_map = config['category_map']
    for r in rows:
        r['mapped_category'] = cat_map.get(r['item_group'] or '') or '(unmapped / excluded)'

    return rows


# ── Public API ────────────────────────────────────────────────────────────────

@frappe.whitelist()
def get_dashboard_data():
    config   = get_config()
    periods  = get_periods()
    snapshot = {}

    for pname, pd in periods.items():
        cy_rows = fetch_rows(pd['cy_start'], pd['cy_end'], config)
        ly_rows = fetch_rows(pd['ly_start'], pd['ly_end'], config)
        snapshot[pname] = {
            'label':    pd['label'],
            'cy_label': pd['cy_label'],
            'ly_label': pd['ly_label'],
            'cy_col':   pd['cy_col'],
            'ly_col':   pd['ly_col'],
            'rows':     aggregate_snapshot(cy_rows, ly_rows),
        }

    return {
        'snapshot': snapshot,
        'as_of': frappe.utils.now_datetime().strftime('%-d %b %Y, %I:%M %p'),
    }


@frappe.whitelist()
def get_brand_data():
    config  = get_config()
    periods = get_periods()
    result  = {}
    for pname in ('daily', 'mtd', 'ytd'):
        pd      = periods[pname]
        cy_rows = fetch_rows(pd['cy_start'], pd['cy_end'], config)
        ly_rows = fetch_rows(pd['ly_start'], pd['ly_end'], config)
        result[pname] = {
            'label':    pd['label'],
            'cy_label': pd['cy_label'],
            'ly_label': pd['ly_label'],
            'cy_col':   pd['cy_col'],
            'ly_col':   pd['ly_col'],
            'rows':     _aggregate_perf(cy_rows, ly_rows, 'brand'),
        }
    return result


@frappe.whitelist()
def get_store_data():
    config  = get_config()
    periods = get_periods()
    result  = {}
    for pname in ('daily', 'mtd', 'ytd'):
        pd      = periods[pname]
        cy_rows = fetch_rows(pd['cy_start'], pd['cy_end'], config)
        ly_rows = fetch_rows(pd['ly_start'], pd['ly_end'], config)
        result[pname] = {
            'label':    pd['label'],
            'cy_label': pd['cy_label'],
            'ly_label': pd['ly_label'],
            'cy_col':   pd['cy_col'],
            'ly_col':   pd['ly_col'],
            'rows':     _aggregate_perf(cy_rows, ly_rows, 'cost_center', b2c_only=False),
        }
    return result


# ── Email ─────────────────────────────────────────────────────────────────────

def _gather_all_data(config):
    periods       = get_periods()
    snapshot_data = {}
    brand_data    = {}
    store_data    = {}

    # Fetch each period once, reuse for all three outputs
    for pname, pd in periods.items():
        cy_rows = fetch_rows(pd['cy_start'], pd['cy_end'], config)
        ly_rows = fetch_rows(pd['ly_start'], pd['ly_end'], config)
        block   = {
            'label':    pd['label'],
            'cy_col':   pd['cy_col'],
            'ly_col':   pd['ly_col'],
            'cy_label': pd['cy_label'],
            'ly_label': pd['ly_label'],
        }
        snapshot_data[pname] = dict(block, rows=aggregate_snapshot(cy_rows, ly_rows))
        if pname in ('daily', 'mtd', 'ytd'):
            brand_data[pname] = dict(block, rows=_aggregate_perf(cy_rows, ly_rows, 'brand'))
            store_data[pname] = dict(block, rows=_aggregate_perf(cy_rows, ly_rows, 'cost_center', b2c_only=False))

    return snapshot_data, brand_data, store_data


def _generate_excel(snapshot_data, brand_data, store_data):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        frappe.throw('openpyxl is not installed. Run: pip install openpyxl')

    wb = openpyxl.Workbook()

    # ── Style helpers ─────────────────────────────────────────────────────────
    def _fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def _font(hex_color, bold=False, size=9):
        return Font(bold=bold, color=hex_color, size=size)

    HDR_FILL = _fill('1E293B')
    HDR_FONT = _font('FFFFFF', bold=True, size=9)
    R_ALIGN  = Alignment(horizontal='right')
    L_ALIGN  = Alignment(horizontal='left')

    # (bg, text, bold, ly_text)  — mirrors Vue CSS exactly
    ROW_PALETTE = {
        'net_revenue':    ('F0F9FF', '0C4A6E', True,  '38BDF8'),
        'b2c':            ('EFF6FF', '1D4ED8', True,  '93C5FD'),
        'b2c_watches':    ('F8FAFF', '475569', False, '94A3B8'),
        'b2c_accessories':('F8FAFF', '475569', False, '94A3B8'),
        'b2c_jewellery':  ('F8FAFF', '475569', False, '94A3B8'),
        'b2c_services':   ('F8FAFF', '475569', False, '94A3B8'),
        'high_value':     ('FAF5FF', '6D28D9', True,  'C4B5FD'),
        'hv_watches':     ('FDFBFF', '475569', False, '94A3B8'),
        'hv_accessories': ('FDFBFF', '475569', False, '94A3B8'),
        'hv_jewellery':   ('FDFBFF', '475569', False, '94A3B8'),
        'hv_services':    ('FDFBFF', '475569', False, '94A3B8'),
    }

    def hdr_cell(cell, val):
        cell.value     = val
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = L_ALIGN if val in ('Row', 'Brand', 'Store', 'Metric') else R_ALIGN

    def title_cell(cell, val):
        cell.value = val
        cell.font  = Font(bold=True, size=10, color='0F172A')

    def pct_fmt(v):
        return None if v is None else round(v, 4)

    def apply_growth_color(cell, v):
        """Green for positive, red for negative, gray for None."""
        if v is None:
            cell.font = Font(color='94A3B8', size=9)
        elif v >= 0:
            cell.font = Font(color='15803D', size=9, bold=True)
        else:
            cell.font = Font(color='B91C1C', size=9, bold=True)

    def style_row(ws, rr, key, num_cols):
        bg, fg, bold, ly_fg = ROW_PALETTE.get(key, ('FFFFFF', '1E293B', False, '64748B'))
        row_fill = _fill(bg)
        for c in range(1, num_cols + 1):
            cell = ws.cell(rr, c)
            cell.fill = row_fill
            # LY columns are 3 and 6 in snapshot table
            if c in (3, 6):
                cell.font = Font(color=ly_fg, size=9, bold=False)
            else:
                cell.font = Font(color=fg, size=9, bold=bold)

    # ── Sheet 1: Snapshot ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Snapshot'
    ws1.column_dimensions['A'].width = 20
    for col in 'BCDEFG': ws1.column_dimensions[col].width = 14

    r = 1
    # KPI summary (MTD values)
    mtd = snapshot_data.get('mtd', {})
    mtd_rows = mtd.get('rows', [])
    def find(key): return next((x for x in mtd_rows if x['key'] == key), {})

    title_cell(ws1.cell(r, 1), 'KPI Summary (MTD)'); r += 1
    for c, h in enumerate(['Metric', 'Value (₹ Cr / Units)', 'Growth %'], 1):
        hdr_cell(ws1.cell(r, c), h)
    r += 1
    net = find('net_revenue'); b2c = find('b2c'); hv = find('high_value')
    for label, val, grwth in [
        ('Net Revenue (₹)',        net.get('cy_rev', 0),  pct_fmt(net.get('growth_rev'))),
        ('Net Qty',                net.get('cy_qty', 0),  pct_fmt(net.get('growth_qty'))),
        ('B2C Revenue (₹)',        b2c.get('cy_rev', 0),  pct_fmt(b2c.get('growth_rev'))),
        ('B2C Qty',                b2c.get('cy_qty', 0),  pct_fmt(b2c.get('growth_qty'))),
        ('High Value Revenue (₹)', hv.get('cy_rev', 0),   pct_fmt(hv.get('growth_rev'))),
        ('High Value Qty',         hv.get('cy_qty', 0),   pct_fmt(hv.get('growth_qty'))),
    ]:
        ws1.cell(r, 1).value = label
        ws1.cell(r, 2).value = round(val, 2)
        ws1.cell(r, 2).number_format = '#,##0'
        ws1.cell(r, 3).value = grwth
        if grwth is not None: ws1.cell(r, 3).number_format = '0.0%'
        r += 1
    r += 1

    def write_snapshot_table(ws, snap, start_row):
        rows   = snap.get('rows', [])
        cy_col = snap.get('cy_col', 'CY')
        ly_col = snap.get('ly_col', 'LY')
        headers = ['Row', f'{cy_col} Rev (₹)', f'{ly_col} Rev (₹)', 'Rev Grwth%',
                   f'{cy_col} Qty', f'{ly_col} Qty', 'Qty Grwth%']
        for c, h in enumerate(headers, 1):
            hdr_cell(ws.cell(start_row, c), h)
        rr = start_row + 1
        for row in rows:
            key    = row['key']
            indent = row['indent']

            # Values
            ws.cell(rr, 1).value = ('    ' if indent else '') + row['label']
            ws.cell(rr, 1).alignment = L_ALIGN
            ws.cell(rr, 2).value = round(row['cy_rev'], 2)
            ws.cell(rr, 2).number_format = '#,##0'
            ws.cell(rr, 2).alignment = R_ALIGN
            ws.cell(rr, 3).value = round(row['ly_rev'], 2)
            ws.cell(rr, 3).number_format = '#,##0'
            ws.cell(rr, 3).alignment = R_ALIGN
            ws.cell(rr, 4).value = pct_fmt(row['growth_rev'])
            ws.cell(rr, 4).alignment = R_ALIGN
            if row['growth_rev'] is not None: ws.cell(rr, 4).number_format = '0.0%'
            ws.cell(rr, 5).value = row['cy_qty']
            ws.cell(rr, 5).alignment = R_ALIGN
            ws.cell(rr, 6).value = row['ly_qty']
            ws.cell(rr, 6).alignment = R_ALIGN
            ws.cell(rr, 7).value = pct_fmt(row['growth_qty'])
            ws.cell(rr, 7).alignment = R_ALIGN
            if row['growth_qty'] is not None: ws.cell(rr, 7).number_format = '0.0%'

            # Row background + text colors from palette
            style_row(ws, rr, key, 7)

            # Override growth columns with green/red
            apply_growth_color(ws.cell(rr, 4), row['growth_rev'])
            apply_growth_color(ws.cell(rr, 7), row['growth_qty'])

            rr += 1
        return rr + 1

    title_cell(ws1.cell(r, 1), f"Daily — {snapshot_data.get('daily',{}).get('cy_col','')} vs {snapshot_data.get('daily',{}).get('ly_col','')}"); r += 1
    r = write_snapshot_table(ws1, snapshot_data.get('daily', {}), r)
    title_cell(ws1.cell(r, 1), f"MTD — {mtd.get('cy_col','')} vs {mtd.get('ly_col','')}"); r += 1
    write_snapshot_table(ws1, mtd, r)

    # ── Sheet 2: Brand Summary ─────────────────────────────────────────────────
    ws2 = wb.create_sheet('Brand Summary')
    ws2.column_dimensions['A'].width = 24
    for col in 'BCDEFG': ws2.column_dimensions[col].width = 14

    def write_perf_table(ws, data, label_header, start_row):
        rows = data.get('rows', [])
        for c, h in enumerate([label_header, 'Sales (₹)', 'Sales Grwth%', 'Units', 'Units Grwth%', 'ASP (₹)', 'ASP Grwth%'], 1):
            hdr_cell(ws.cell(start_row, c), h)
        rr = start_row + 1
        for i, row in enumerate(rows):
            alt_fill = _fill('F8FAFC') if i % 2 == 1 else _fill('FFFFFF')
            ws.cell(rr, 1).value = row['name']
            ws.cell(rr, 1).font  = Font(color='0F172A', bold=True, size=9)
            ws.cell(rr, 1).fill  = alt_fill
            ws.cell(rr, 1).alignment = L_ALIGN

            ws.cell(rr, 2).value          = round(row['cy_sales'], 2)
            ws.cell(rr, 2).number_format  = '#,##0'
            ws.cell(rr, 2).font           = Font(color='1E293B', size=9)
            ws.cell(rr, 2).fill           = alt_fill
            ws.cell(rr, 2).alignment      = R_ALIGN

            ws.cell(rr, 3).value          = pct_fmt(row['growth_sales'])
            if row['growth_sales'] is not None: ws.cell(rr, 3).number_format = '0.0%'
            ws.cell(rr, 3).fill           = alt_fill
            ws.cell(rr, 3).alignment      = R_ALIGN
            apply_growth_color(ws.cell(rr, 3), row['growth_sales'])

            ws.cell(rr, 4).value          = row['cy_units']
            ws.cell(rr, 4).font           = Font(color='1E293B', size=9)
            ws.cell(rr, 4).fill           = alt_fill
            ws.cell(rr, 4).alignment      = R_ALIGN

            ws.cell(rr, 5).value          = pct_fmt(row['growth_units'])
            if row['growth_units'] is not None: ws.cell(rr, 5).number_format = '0.0%'
            ws.cell(rr, 5).fill           = alt_fill
            ws.cell(rr, 5).alignment      = R_ALIGN
            apply_growth_color(ws.cell(rr, 5), row['growth_units'])

            ws.cell(rr, 6).value          = round(row['cy_asp'], 2) if row['cy_asp'] else 0
            ws.cell(rr, 6).number_format  = '#,##0'
            ws.cell(rr, 6).font           = Font(color='1E293B', size=9)
            ws.cell(rr, 6).fill           = alt_fill
            ws.cell(rr, 6).alignment      = R_ALIGN

            ws.cell(rr, 7).value          = pct_fmt(row['growth_asp'])
            if row['growth_asp'] is not None: ws.cell(rr, 7).number_format = '0.0%'
            ws.cell(rr, 7).fill           = alt_fill
            ws.cell(rr, 7).alignment      = R_ALIGN
            apply_growth_color(ws.cell(rr, 7), row['growth_asp'])

            rr += 1
        return rr + 1

    r2 = 1
    for pname in ('daily', 'mtd', 'ytd'):
        pd = brand_data.get(pname, {})
        title_cell(ws2.cell(r2, 1), f"{pd.get('label', pname.upper())} — {pd.get('cy_col','')} vs {pd.get('ly_col','')}"); r2 += 1
        r2 = write_perf_table(ws2, pd, 'Brand', r2)

    # ── Sheet 3: Store Summary ─────────────────────────────────────────────────
    ws3 = wb.create_sheet('Store Summary')
    ws3.column_dimensions['A'].width = 32
    for col in 'BCDEFG': ws3.column_dimensions[col].width = 14

    r3 = 1
    for pname in ('daily', 'mtd', 'ytd'):
        pd = store_data.get(pname, {})
        title_cell(ws3.cell(r3, 1), f"{pd.get('label', pname.upper())} — {pd.get('cy_col','')} vs {pd.get('ly_col','')}"); r3 += 1
        r3 = write_perf_table(ws3, pd, 'Store', r3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _email_html(snapshot_data):
    daily  = snapshot_data.get('daily', {})
    rows   = daily.get('rows', [])
    cy_col = daily.get('cy_col', 'CY')
    ly_col = daily.get('ly_col', 'LY')

    def cr(v):  return '—' if v is None else f'₹{v/1e7:.2f} Cr'
    def pct(v): return '—' if v is None else f"{'▲' if v >= 0 else '▼'} {abs(v*100):.1f}%"
    def clr(v): return '#16a34a' if (v or 0) >= 0 else '#dc2626'

    body = ''.join(f"""
      <tr>
        <td style="padding:6px 12px;{'padding-left:24px;color:#64748b' if r['indent'] else 'font-weight:600'}">{r['label']}</td>
        <td style="padding:6px 12px;text-align:right">{cr(r['cy_rev'])}</td>
        <td style="padding:6px 12px;text-align:right;color:#64748b">{cr(r['ly_rev'])}</td>
        <td style="padding:6px 12px;text-align:right;color:{clr(r['growth_rev'])}">{pct(r['growth_rev'])}</td>
        <td style="padding:6px 12px;text-align:right">{r['cy_qty']:,}</td>
        <td style="padding:6px 12px;text-align:right;color:{clr(r['growth_qty'])}">{pct(r['growth_qty'])}</td>
      </tr>""" for r in rows)

    return f"""<html><body style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;color:#1e293b;max-width:700px;margin:auto">
<h2 style="color:#0f172a">AOT Sales Dashboard — {date.today().strftime('%-d %b %Y')}</h2>
<h3 style="color:#334155;margin-bottom:4px">Daily Snapshot</h3>
<p style="color:#64748b;margin-top:0">{daily.get('cy_label','')} vs {daily.get('ly_label','')}</p>
<table cellspacing="0" style="border-collapse:collapse;font-size:13px;width:100%">
  <thead><tr style="background:#1e293b;color:#fff">
    <th style="padding:8px 12px;text-align:left">Row</th>
    <th style="padding:8px 12px;text-align:right">{cy_col}</th>
    <th style="padding:8px 12px;text-align:right">{ly_col}</th>
    <th style="padding:8px 12px;text-align:right">Rev Grwth</th>
    <th style="padding:8px 12px;text-align:right">Qty CY</th>
    <th style="padding:8px 12px;text-align:right">Qty Grwth</th>
  </tr></thead>
  <tbody>{body}</tbody>
</table>
<p style="color:#94a3b8;font-size:12px;margin-top:20px">Full data attached as Excel (3 sheets: Snapshot, Brand Summary, Store Summary).</p>
</body></html>"""


@frappe.whitelist()
def download_excel():
    config = get_config()
    snapshot_data, brand_data, store_data = _gather_all_data(config)
    xlsx = _generate_excel(snapshot_data, brand_data, store_data)
    fname = f'AOT_Dashboard_{date.today():%Y%m%d}.xlsx'
    frappe.response['filename']     = fname
    frappe.response['filecontent']  = xlsx
    frappe.response['type']         = 'download'
    frappe.response['content_type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@frappe.whitelist()
def send_test_email(recipient=None):
    if not recipient:
        frappe.throw('Recipient email is required')
    config = get_config()
    doc    = frappe.get_single('AOT Dashboard Config')
    snapshot_data, brand_data, store_data = _gather_all_data(config)
    xlsx   = _generate_excel(snapshot_data, brand_data, store_data)
    subject = (doc.email_subject or 'AOT Sales Dashboard – {date}').format(
        date=date.today().strftime('%-d %b %Y'))
    frappe.sendmail(
        recipients=[recipient],
        subject=f'[TEST] {subject}',
        message=_email_html(snapshot_data),
        attachments=[{'fname': f'AOT_Dashboard_{date.today():%Y%m%d}.xlsx', 'fcontent': xlsx}],
    )
    return {'status': 'sent', 'recipient': recipient}


def check_and_send_dashboard_email():
    """
    Hourly scheduler hook.
    Fires send_dashboard_email() when the current hour matches the
    configured send_time in AOT Dashboard Config.
    A Redis key prevents double-sending within the same calendar day.
    """
    try:
        doc = frappe.get_single('AOT Dashboard Config')

        if not doc.send_time:
            return
        recipients = [r.email for r in (doc.email_recipients or []) if r.email]
        if not recipients:
            return

        from datetime import datetime
        now       = datetime.now()
        send_hour = int(str(doc.send_time).split(':')[0])

        if now.hour != send_hour:
            return

        # Guard: only one send per calendar day
        cache_key = f'aot_dashboard_email_sent_{now.date()}'
        if frappe.cache().get_value(cache_key):
            return

        send_dashboard_email()
        # Expires after 25 h so the key clears before the next day's window
        frappe.cache().set_value(cache_key, 1, expires_in_sec=90000)

    except Exception:
        frappe.log_error(frappe.get_traceback(), 'AOT Dashboard Email Scheduler')


@frappe.whitelist()
def send_dashboard_email():
    """Called by scheduler — sends to all configured recipients."""
    config     = get_config()
    doc        = frappe.get_single('AOT Dashboard Config')
    recipients = [r.email for r in (doc.email_recipients or []) if r.email]
    if not recipients:
        return
    snapshot_data, brand_data, store_data = _gather_all_data(config)
    xlsx   = _generate_excel(snapshot_data, brand_data, store_data)
    subject = (doc.email_subject or 'AOT Sales Dashboard – {date}').format(
        date=date.today().strftime('%-d %b %Y'))
    frappe.sendmail(
        recipients=recipients,
        subject=subject,
        message=_email_html(snapshot_data),
        attachments=[{'fname': f'AOT_Dashboard_{date.today():%Y%m%d}.xlsx', 'fcontent': xlsx}],
    )
